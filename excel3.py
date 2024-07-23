import openpyxl

book = openpyxl.load_workbook("C:\\Users\\nevai\\AppData\\Local\\Programs\\Python\\Python312\\Lib\\site-packages\\SecondBook.xlsx")
sheet = book.active

rows = sheet.rows
book.close()

values = []
for row in rows:
    for cell in row:
        value = (cell.value)
        if (cell.column) == 2:
            value = value * 10
        values.append(value)

book = openpyxl.Workbook()
sheet = book.active

index = 0
rw = 1
col = 1
for val in values:
    sheet.cell(row = rw, column = col).value = val
    index += 1
    col += 1
    if index == 3:
        rw +=1
        col = 1
        index = 0

book.save("C:\\Users\\nevai\\AppData\\Local\\Programs\\Python\\Python312\\Lib\\site-packages\\ThirdBook.xlsx")
